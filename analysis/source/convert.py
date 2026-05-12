"""Dumps plan.xlsx into one markdown per sheet for downstream analysis."""

from pathlib import Path

import openpyxl

SRC = Path(__file__).parent / "plan.xlsx"
OUT = Path(__file__).parent / "sheets"
OUT.mkdir(exist_ok=True)

DAYS = ("MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY")


def merged_skip(ws):
    skip = set()
    for rng in ws.merged_cells.ranges:
        for rr in range(rng.min_row, rng.max_row + 1):
            for cc in range(rng.min_col, rng.max_col + 1):
                if (rr, cc) != (rng.min_row, rng.min_col):
                    skip.add((rr, cc))
    return skip


def cell_text(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return None
    text = str(v).strip()
    return text or None


def dump_week(ws):
    skip = merged_skip(ws)
    day_cols = []
    for c in range(1, ws.max_column + 1):
        v = cell_text(ws, 1, c)
        if v and v.upper() in DAYS:
            day_cols.append((v.upper(), c))

    lines = []
    for day, col in day_cols:
        lines.append(f"## {day}")
        lines.append("")
        had_content = False
        for r in range(2, ws.max_row + 1):
            if (r, col) in skip:
                continue
            t = cell_text(ws, r, col)
            if t is None:
                continue
            lines.append(f"- (r{r}) {t}")
            had_content = True
        if not had_content:
            lines.append("_(empty)_")
        lines.append("")
    return "\n".join(lines)


def dump_quick_search(ws):
    lines = ["# Quick Search", ""]
    skip = merged_skip(ws)
    for r in range(1, ws.max_row + 1):
        row_items = []
        for c in range(1, ws.max_column + 1):
            if (r, c) in skip:
                continue
            t = cell_text(ws, r, c)
            if t is None:
                continue
            row_items.append((c, t))
        if not row_items:
            continue
        for c, t in row_items:
            lines.append(f"- (r{r}c{c}) {t}")
    return "\n".join(lines)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    qs = wb["Quick Search"]
    (OUT / "quick-search.md").write_text(dump_quick_search(qs) + "\n", encoding="utf-8")

    weeks = [n for n in wb.sheetnames if n != "Quick Search"]
    for idx, name in enumerate(weeks, start=1):
        ws = wb[name]
        title = name.strip()
        body = f"# {title}\n\n{dump_week(ws)}\n"
        (OUT / f"sheet-{idx:02d}.md").write_text(body, encoding="utf-8")

    print(f"wrote {len(weeks) + 1} files to {OUT}")


if __name__ == "__main__":
    main()
